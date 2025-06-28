import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

ORDERS_FILE = os.path.join(os.path.dirname(__file__), 'orders.xlsx')

HEADERS = ['Order ID', 'User ID', 'Delivery Info', 'Created At', 'Status', 'Product Name', 'Quantity', 'Price']

def append_order_to_excel(order_id, user_id, delivery_info, created_at, status, order_items):
    if not os.path.exists(ORDERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(ORDERS_FILE)

    wb = load_workbook(ORDERS_FILE)
    ws = wb.active

    created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S')

    # Добавляем каждую позицию заказа как отдельную строку
    for item in order_items:
        ws.append([
            order_id, user_id, delivery_info, created_at_str, status,
            item['product_name'], item['quantity'], item['product_price']
        ])

    wb.save(ORDERS_FILE)
